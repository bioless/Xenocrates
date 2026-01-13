#!/usr/bin/env python3
"""
Script to create test files for Excel and JSON format testing.
"""

import json
from openpyxl import Workbook

# Test data (subset of test-data-basic.tsv)
test_entries = [
    {
        "Title": "AES Encryption",
        "Book": "SEC401",
        "Page": "142",
        "Description": "Advanced Encryption Standard - 128/192/256 bit symmetric cipher"
    },
    {
        "Title": "Apache Web Server",
        "Book": "SEC504",
        "Page": "87",
        "Description": "Popular open-source HTTP server software"
    },
    {
        "Title": "API Security",
        "Book": "SEC522",
        "Page": "201",
        "Description": "Application Programming Interface security best practices"
    },
    {
        "Title": "Base64 Encoding",
        "Book": "FOR500",
        "Page": "93",
        "Description": "Binary-to-text encoding scheme using 64 ASCII characters"
    },
    {
        "Title": "Bash Scripting",
        "Book": "SEC505",
        "Page": "156",
        "Description": "Unix shell scripting for automation and incident response"
    },
    {
        "Title": "Boolean Logic",
        "Book": "FOR508",
        "Page": "44",
        "Description": "AND, OR, NOT operations in digital forensics"
    },
    {
        "Title": "Cross-Site Scripting",
        "Book": "SEC542",
        "Page": "78",
        "Description": "XSS attack - injection of malicious scripts into web pages"
    },
    {
        "Title": "CSRF Tokens",
        "Book": "SEC542",
        "Page": "82",
        "Description": "Cross-Site Request Forgery protection mechanism"
    },
    {
        "Title": "CVE Database",
        "Book": "SEC504",
        "Page": "34",
        "Description": "Common Vulnerabilities and Exposures - public database"
    },
    {
        "Title": "DNS Tunneling",
        "Book": "SEC503",
        "Page": "167",
        "Description": "Covert channel using DNS protocol for data exfiltration"
    },
]

# GSE test data with Course column
gse_entries = [
    {
        "Title": "Kerberos Authentication",
        "Book": "SEC505",
        "Page": "201",
        "Description": "Network authentication protocol using tickets and symmetric encryption",
        "Course": "SEC575"
    },
    {
        "Title": "PowerShell Remoting",
        "Book": "SEC505",
        "Page": "145",
        "Description": "Remote command execution using Windows PowerShell",
        "Course": "SEC505"
    },
    {
        "Title": "IDS vs IPS",
        "Book": "SEC503",
        "Page": "89",
        "Description": "Intrusion Detection vs Intrusion Prevention Systems comparison",
        "Course": "SEC503"
    },
]

# Create Excel file - standard format
print("Creating test-data-excel.xlsx...")
wb = Workbook()
ws = wb.active
ws.title = "Index"

# Write headers
ws.append(["Title", "Book", "Page", "Description"])

# Write data
for entry in test_entries:
    ws.append([entry["Title"], entry["Book"], entry["Page"], entry["Description"]])

wb.save("test-data-excel.xlsx")
print(f"  ✓ Created with {len(test_entries)} entries")

# Create Excel file - GSE format with Course column
print("Creating test-data-excel-gse.xlsx...")
wb_gse = Workbook()
ws_gse = wb_gse.active
ws_gse.title = "Index"

# Write headers
ws_gse.append(["Title", "Book", "Page", "Description", "Course"])

# Write data
for entry in gse_entries:
    ws_gse.append([entry["Title"], entry["Book"], entry["Page"], entry["Description"], entry["Course"]])

wb_gse.save("test-data-excel-gse.xlsx")
print(f"  ✓ Created with {len(gse_entries)} entries (GSE mode)")

# Create Excel file - lowercase columns
print("Creating test-excel-lowercase-columns.xlsx...")
wb_lower = Workbook()
ws_lower = wb_lower.active
ws_lower.title = "Index"

# Write headers in lowercase
ws_lower.append(["title", "book", "page", "description"])

# Write first 3 entries
for entry in test_entries[:3]:
    ws_lower.append([entry["Title"], entry["Book"], entry["Page"], entry["Description"]])

wb_lower.save("test-excel-lowercase-columns.xlsx")
print(f"  ✓ Created with lowercase column names")

# Create JSON file - wrapped format
print("Creating test-data.json...")
json_data = {"entries": test_entries}
with open("test-data.json", "w", encoding="utf-8") as f:
    json.dump(json_data, f, indent=2, ensure_ascii=False)
print(f"  ✓ Created with {len(test_entries)} entries (wrapped format)")

# Create JSON file - direct array format
print("Creating test-data-json-direct.json...")
with open("test-data-json-direct.json", "w", encoding="utf-8") as f:
    json.dump(test_entries, f, indent=2, ensure_ascii=False)
print(f"  ✓ Created (direct array format)")

# Create JSON file - GSE format
print("Creating test-data-json-gse.json...")
json_gse = {"entries": gse_entries}
with open("test-data-json-gse.json", "w", encoding="utf-8") as f:
    json.dump(json_gse, f, indent=2, ensure_ascii=False)
print(f"  ✓ Created with {len(gse_entries)} entries (GSE mode)")

print("\n✅ All test files created successfully!")
